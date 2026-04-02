#!/usr/bin/env python3
"""Excel (.xlsx) → Markdown 変換スクリプト

openpyxl でシートごとにデータを読み取り、Markdownテーブル形式で出力する。

Usage:
    python3 xlsx2md.py <input.xlsx> [--output <output.md>] [--max-rows 500] [--format md|csv]
"""

import argparse
import csv
import sys
from io import StringIO
from pathlib import Path


def cell_value_to_str(cell) -> str:
    """セル値を文字列に変換する。"""
    if cell.value is None:
        return ""
    return str(cell.value).replace("\n", " ").replace("|", "\\|").strip()


def sheet_to_md_table(ws, max_rows: int) -> tuple[list[str], dict]:
    """ワークシートを Markdown テーブルに変換する。"""
    lines: list[str] = []
    stats = {"rows": 0, "cols": 0, "truncated": False}

    # データ範囲を取得
    rows = list(ws.iter_rows())
    if not rows:
        lines.append("*（空のシート）*")
        return lines, stats

    stats["cols"] = ws.max_column
    total_rows = len(rows)

    # 結合セルの情報を取得
    merged_cells = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        value = ws.cell(row=min_row, column=min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_cells[(row, col)] = value if value is not None else ""

    # 行数制限
    display_rows = rows[:max_rows] if total_rows > max_rows else rows
    if total_rows > max_rows:
        stats["truncated"] = True

    for i, row in enumerate(display_rows):
        stats["rows"] += 1
        cells: list[str] = []
        for j, cell in enumerate(row, 1):
            # 結合セルの場合は結合元の値を使用
            key = (cell.row, cell.column)
            if key in merged_cells:
                cells.append(str(merged_cells[key]).replace("\n", " ").replace("|", "\\|").strip())
            else:
                cells.append(cell_value_to_str(cell))

        lines.append("| " + " | ".join(cells) + " |")

        # 1行目の後にセパレータを挿入
        if i == 0:
            lines.append("| " + " | ".join(["---"] * len(cells)) + " |")

    if stats["truncated"]:
        lines.append("")
        lines.append(f"> ⚠️ {total_rows}行中 {max_rows}行まで表示（`--max-rows` で変更可能）")

    return lines, stats


def sheet_to_csv(ws, max_rows: int) -> tuple[str, dict]:
    """ワークシートを CSV 文字列に変換する。"""
    stats = {"rows": 0, "cols": ws.max_column, "truncated": False}
    output = StringIO()
    writer = csv.writer(output)

    rows = list(ws.iter_rows())
    total_rows = len(rows)
    display_rows = rows[:max_rows] if total_rows > max_rows else rows
    if total_rows > max_rows:
        stats["truncated"] = True

    for row in display_rows:
        stats["rows"] += 1
        writer.writerow([cell_value_to_str(cell) for cell in row])

    return output.getvalue(), stats


def convert_xlsx(input_path: Path, output_path: Path, max_rows: int, fmt: str) -> dict:
    """xlsx を Markdown または CSV に変換する。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("エラー: openpyxl が未インストールです。", file=sys.stderr)
        print("  pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(str(input_path), read_only=False, data_only=True)
    all_stats = {"sheets": 0, "total_rows": 0}

    if fmt == "csv":
        # CSV: シートごとに別ファイル
        for ws in wb.worksheets:
            all_stats["sheets"] += 1
            sheet_output = output_path.parent / f"{output_path.stem}_{ws.title}.csv"
            csv_content, stats = sheet_to_csv(ws, max_rows)
            sheet_output.write_text(csv_content, encoding="utf-8")
            all_stats["total_rows"] += stats["rows"]
            print(f"  シート「{ws.title}」→ {sheet_output.name} ({stats['rows']}行)")
            if stats["truncated"]:
                print(f"    ⚠️ 行数制限により切り詰め（上限: {max_rows}行）")
    else:
        # Markdown: 1ファイルにまとめる
        lines: list[str] = []
        lines.append(f"# {input_path.stem}")
        lines.append("")

        for ws in wb.worksheets:
            all_stats["sheets"] += 1
            lines.append(f"## {ws.title}")
            lines.append("")
            table_lines, stats = sheet_to_md_table(ws, max_rows)
            lines.extend(table_lines)
            lines.append("")
            all_stats["total_rows"] += stats["rows"]

        output_path.write_text("\n".join(lines), encoding="utf-8")

    wb.close()
    return all_stats


def main():
    parser = argparse.ArgumentParser(description="Excel (.xlsx) → Markdown/CSV 変換")
    parser.add_argument("input", help="入力 .xlsx ファイルパス")
    parser.add_argument("--output", "-o", help="出力ファイルパス（省略時: 入力と同名.md/.csv）")
    parser.add_argument("--max-rows", type=int, default=500, help="シートあたりの最大行数（デフォルト: 500）")
    parser.add_argument("--format", choices=["md", "csv"], default="md", help="出力形式（デフォルト: md）")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if not input_path.exists():
        print(f"エラー: ファイルが見つかりません: {input_path}", file=sys.stderr)
        sys.exit(1)
    if input_path.suffix.lower() != ".xlsx":
        print(f"エラー: .xlsx ファイルを指定してください: {input_path}", file=sys.stderr)
        sys.exit(1)

    default_ext = ".csv" if args.format == "csv" else ".md"
    output_path = Path(args.output).resolve() if args.output else input_path.with_suffix(default_ext)

    stats = convert_xlsx(input_path, output_path, args.max_rows, args.format)

    print(f"変換完了: {input_path.name} → {output_path.name}")
    print(f"  シート: {stats['sheets']}件  合計行数: {stats['total_rows']}行")


if __name__ == "__main__":
    main()
